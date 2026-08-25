"""Recipe PDF Loader Module for RestaurantOS.

Parses recipe structures from ``productosestructura.frx.pdf`` (SoftRestaurant export),
synchronizes required base supplies and products, and populates canonical versioned recipes
into PostgreSQL / SQLite database with accurate theoretical costing.
"""

from __future__ import annotations

# ruff: noqa: E501
import json
import os
import re
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

import sqlalchemy as sa
from sqlalchemy.orm import Session

from . import models
from .operations import (
    ORGANIZATION_ID,
    _now,
    _quantity,
)
from .real_catalog_loader import load_real_catalog_from_excels


def _uid() -> str:
    return str(uuid4())


def parse_pdf_recipe_catalog(pdf_path: str = "productosestructura.frx.pdf") -> list[dict[str, Any]]:
    """Extracts all product recipes and component lists from the PDF structure report or bundled JSON."""
    # 1. Check for bundled pre-parsed catalog JSON
    possible_json_paths = [
        Path(__file__).resolve().parent / "data" / "recipes_catalog_data.json",
        Path("restaurant_os/data/recipes_catalog_data.json"),
        Path("apps/api/restaurant_os/data/recipes_catalog_data.json"),
        Path("/app/apps/api/restaurant_os/data/recipes_catalog_data.json"),
    ]
    for jp in possible_json_paths:
        if jp.exists():
            with open(jp, "r", encoding="utf-8") as f:
                return json.load(f)

    # 2. Parse from PDF if file exists
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"Recipe report PDF not found at {pdf_path}")

    try:
        import pymupdf as fitz
    except ImportError:
        try:
            import fitz
        except ImportError:
            raise ImportError("PyMuPDF is required to parse PDF recipe catalogs.")

    doc = fitz.open(pdf_path)
    sku_pattern = re.compile(r"^([A-Za-z0-9]+)\s*-\s*(.+)$")
    possible_units = {
        "KILO", "KILOS", "KG", "GR", "GRAMO", "GRAMOS",
        "LITRO", "LITROS", "L", "LT", "LTS", "ML", "MILILITRO", "MILILITROS",
        "PZA", "PIEZA", "PIEZAS", "PZ", "PORCION", "PORCIONES", "PAQUETE"
    }

    parsed_recipes: list[dict[str, Any]] = []
    current_group = "GENERAL"
    current_prod: dict[str, Any] | None = None

    for page_idx in range(len(doc)):
        page = doc[page_idx]
        blocks = page.get_text("blocks")
        for b in blocks:
            text = b[4].strip()
            lines = [l.strip() for l in text.split("\n") if l.strip()]
            if not lines:
                continue

            if lines[0].startswith("GRUPO:"):
                current_group = lines[0].replace("GRUPO:", "").strip()
                continue

            if "KIWI COMIDA NATURAL" in text or "RECETA DE PRODUCTOS" in text or "SoftRestaurant" in text:
                continue
            if lines[0] in ("PRODUCTO", "COSTO PROMEDIO", "COSTO", "INSUMO", "PRECIO DE VENTA"):
                continue
            if len(lines) == 1 and lines[0].startswith("$") and current_prod is not None:
                continue

            m = sku_pattern.match(lines[0])
            if m:
                sku = m.group(1).strip()
                name = m.group(2).strip()

                is_insumo = False
                for l in lines[1:]:
                    if l.upper() in possible_units:
                        is_insumo = True
                        break

                if not is_insumo and (any("%" in l for l in lines) or (len(lines) >= 2 and lines[1].startswith("$"))):
                    price_str = None
                    if len(lines) > 1 and lines[1].startswith("$"):
                        price_str = lines[1].replace("$", "").replace(",", "").strip()
                    current_prod = {
                        "group": current_group,
                        "sku": sku,
                        "name": name,
                        "page": page_idx + 1,
                        "price": price_str,
                        "components": [],
                    }
                    parsed_recipes.append(current_prod)
                    continue

                if is_insumo and current_prod is not None:
                    unit_cost = None
                    qty = None
                    unit = None
                    total_cost = None
                    for l in lines[1:]:
                        if l.startswith("$"):
                            val = l.replace("$", "").replace(",", "").strip()
                            if unit_cost is None:
                                unit_cost = val
                            else:
                                total_cost = val
                        elif l.upper() in possible_units:
                            unit = l.upper()
                        else:
                            try:
                                float(l.replace(",", ""))
                                qty = l.replace(",", "").strip()
                            except ValueError:
                                pass

                    current_prod["components"].append({
                        "insumo_sku": sku,
                        "insumo_name": name,
                        "unit_cost": unit_cost,
                        "quantity": qty,
                        "unit": unit,
                        "total_cost": total_cost,
                    })

    return parsed_recipes


def _ensure_missing_supplies_and_products(
    session: Session, organization_id: str, excel_dir: str = "."
) -> None:
    """Ensures essential Coffee, Matcha and Extra Protein supplies and products are present."""
    now = _now()
    
    # 1. Base Inventory Units
    kilo_unit = session.execute(
        sa.select(models.inventory_units.c.id).where(
            models.inventory_units.c.organization_id == organization_id,
            models.inventory_units.c.code == "KG",
        )
    ).scalars().first()
    if not kilo_unit:
        kilo_unit = session.execute(
            sa.select(models.inventory_units.c.id).where(
                models.inventory_units.c.organization_id == organization_id
            )
        ).scalars().first()

    # 2. Missing Insumos: CAFE MOLIDO (1026), MACCHA (1027), PROTEINA (1028)
    extra_insumos = [
        {"sku": "1026", "name": "CAFE MOLIDO", "avg_cost": Decimal("88.0473"), "last_cost": Decimal("88.0473"), "tax": Decimal("16.0")},
        {"sku": "1027", "name": "MACCHA", "avg_cost": Decimal("0.0"), "last_cost": Decimal("0.0"), "tax": Decimal("0.0")},
        {"sku": "1028", "name": "PROTEINA", "avg_cost": Decimal("0.0"), "last_cost": Decimal("0.0"), "tax": Decimal("0.0")},
    ]
    for ins in extra_insumos:
        existing = session.execute(
            sa.select(models.inventory_items.c.id).where(
                models.inventory_items.c.organization_id == organization_id,
                models.inventory_items.c.sku == ins["sku"],
            )
        ).scalars().first()
        if not existing:
            session.execute(
                models.inventory_items.insert().values(
                    id=_uid(),
                    organization_id=organization_id,
                    sku=ins["sku"],
                    name=ins["name"],
                    base_unit_id=kilo_unit,
                    category_name="ABARROTE",
                    item_type="ingredient",
                    catalog_scope="organization",
                    source_branch_id=None,
                    status="active",
                    created_at=now,
                    updated_at=now,
                )
            )

    # 3. Ensure all categories and products from PRODUCTOS.XLS exist
    prod_path = os.path.join(excel_dir, "PRODUCTOS.XLS")
    if os.path.exists(prod_path):
        import io
        import openpyxl
        with open(prod_path, "rb") as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws_prod = wb["curtoexcel_1"]
        
        category_map: dict[str, str] = {}
        for r in range(6, ws_prod.max_row + 1):
            raw_sku = ws_prod.cell(r, 1).value
            desc = ws_prod.cell(r, 2).value
            grupo = ws_prod.cell(r, 4).value
            precio = ws_prod.cell(r, 8).value
            if not raw_sku or not desc:
                continue

            clave = str(raw_sku).strip().lstrip("'")
            name = str(desc).strip()
            group_name = str(grupo).strip() if grupo else "GENERAL"
            price_val = float(precio) if precio is not None else 0.0
            price_cents = int(round(price_val * 100))

            if group_name not in category_map:
                existing_cat = session.execute(
                    sa.select(models.product_categories.c.id).where(
                        models.product_categories.c.organization_id == organization_id,
                        models.product_categories.c.name == group_name,
                    )
                ).scalars().first()
                if existing_cat:
                    cat_id = str(existing_cat)
                else:
                    cat_id = f"cat-{group_name.lower().replace(' ', '-').replace('/', '-')[:30]}"
                    session.execute(
                        models.product_categories.insert().values(
                            id=cat_id,
                            organization_id=organization_id,
                            name=group_name[:120],
                            display_order=len(category_map) + 1,
                            status="active",
                            created_at=now,
                            updated_at=now,
                        )
                    )
                category_map[group_name] = cat_id
            else:
                cat_id = category_map[group_name]

            sku = f"PROD-{clave}"
            station = "cocina" if any(k in group_name for k in ["ENSALADA", "SANDWICH", "BAGUETTE", "FOCACCIA", "CUERNITO", "QUESADILLA", "COMBOS", "OMELETTE"]) else "barra"

            existing_p = session.execute(
                sa.select(models.products.c.id).where(
                    models.products.c.organization_id == organization_id,
                    models.products.c.sku.in_([clave, sku, clave.zfill(5)]),
                )
            ).scalars().first()

            if not existing_p:
                prod_id = f"prod-{clave.lower()}"
                session.execute(
                    models.products.insert().values(
                        id=prod_id,
                        organization_id=organization_id,
                        category_id=cat_id,
                        name=name[:160],
                        sku=sku[:64],
                        description=f"{name} - {group_name}",
                        station=station,
                        status="active",
                        catalog_scope="organization",
                        source_branch_id=None,
                        created_at=now,
                        updated_at=now,
                    )
                )
                session.execute(
                    models.price_versions.insert().values(
                        id=_uid(),
                        organization_id=organization_id,
                        product_id=prod_id,
                        price_cents=price_cents,
                        currency="MXN",
                        valid_from=now,
                        valid_to=None,
                        created_at=now,
                    )
                )

    # 4. Missing Categories for Café y Matcha
    cafe_cat = session.execute(
        sa.select(models.product_categories.c.id).where(
            models.product_categories.c.organization_id == organization_id,
            models.product_categories.c.name == "CAFE Y MACCHA",
        )
    ).scalars().first()
    if not cafe_cat:
        cafe_cat = "cat-cafe-y-maccha"
        session.execute(
            models.product_categories.insert().values(
                id=cafe_cat,
                organization_id=organization_id,
                name="CAFE Y MACCHA",
                display_order=20,
                status="active",
                created_at=now,
                updated_at=now,
            )
        )

    extra_cat = session.execute(
        sa.select(models.product_categories.c.id).where(
            models.product_categories.c.organization_id == organization_id,
            models.product_categories.c.name == "INGREDIENTE EXTRA",
        )
    ).scalars().first()
    if not extra_cat:
        extra_cat = "cat-ingrediente-extra"
        session.execute(
            models.product_categories.insert().values(
                id=extra_cat,
                organization_id=organization_id,
                name="INGREDIENTE EXTRA",
                display_order=21,
                status="active",
                created_at=now,
                updated_at=now,
            )
        )

    # 5. Missing Products
    extra_products = [
        {"sku": "24001", "name": "CAFE SOLO", "category_id": cafe_cat, "price_cents": 5000, "station": "barra"},
        {"sku": "24002", "name": "CAFE SOLO FRESH", "category_id": cafe_cat, "price_cents": 5500, "station": "barra"},
        {"sku": "24003", "name": "CAFE NARANJA", "category_id": cafe_cat, "price_cents": 7500, "station": "barra"},
        {"sku": "24004", "name": "KIWI LATTE", "category_id": cafe_cat, "price_cents": 7000, "station": "barra"},
        {"sku": "24005", "name": "KIWI LATTE FRESH", "category_id": cafe_cat, "price_cents": 7500, "station": "barra"},
        {"sku": "24006", "name": "MACCHA SHIRU", "category_id": cafe_cat, "price_cents": 10000, "station": "barra"},
        {"sku": "24007", "name": "MACCHA PINKU (CON FRESA)", "category_id": cafe_cat, "price_cents": 11000, "station": "barra"},
        {"sku": "11057", "name": "PROTEINA", "category_id": extra_cat, "price_cents": 3000, "station": "barra"},
    ]
    for p in extra_products:
        existing = session.execute(
            sa.select(models.products.c.id).where(
                models.products.c.organization_id == organization_id,
                models.products.c.sku.in_([p["sku"], f"PROD-{p['sku']}"]),
            )
        ).scalars().first()
        if not existing:
            prod_id = f"prod-{p['sku'].lower()}"
            session.execute(
                models.products.insert().values(
                    id=prod_id,
                    organization_id=organization_id,
                    category_id=p["category_id"],
                    name=p["name"],
                    sku=f"PROD-{p['sku']}",
                    description=p["name"],
                    station=p["station"],
                    status="active",
                    catalog_scope="organization",
                    source_branch_id=None,
                    created_at=now,
                    updated_at=now,
                )
            )
            session.execute(
                models.price_versions.insert().values(
                    id=_uid(),
                    organization_id=organization_id,
                    product_id=prod_id,
                    price_cents=p["price_cents"],
                    currency="MXN",
                    valid_from=now,
                    valid_to=None,
                    created_at=now,
                )
            )


def load_recipes_from_pdf(
    session: Session,
    pdf_path: str = "productosestructura.frx.pdf",
    excel_dir: str = ".",
    organization_id: str = ORGANIZATION_ID,
) -> dict[str, Any]:
    """Imports and configures all valid recipes from the PDF report into the database."""
    now = _now()

    # Step 1: Ensure base excel catalogs and extra items are loaded
    load_real_catalog_from_excels(session, excel_dir=excel_dir, organization_id=organization_id, import_customers=False)
    _ensure_missing_supplies_and_products(session, organization_id, excel_dir=excel_dir)
    session.flush()

    # Step 2: Parse raw recipes from PDF
    parsed_recipes = parse_pdf_recipe_catalog(pdf_path)

    # Step 3: Build index mappings
    all_prods = session.execute(
        sa.select(models.products.c.id, models.products.c.sku, models.products.c.name).where(
            models.products.c.organization_id == organization_id
        )
    ).fetchall()
    prod_map: dict[str, str] = {}
    for p in all_prods:
        clean_sku = p.sku.replace("PROD-", "").strip().lstrip("'")
        prod_map[clean_sku] = p.id
        if clean_sku.isdigit():
            prod_map[str(int(clean_sku))] = p.id
            prod_map[clean_sku.zfill(5)] = p.id

    all_items = session.execute(
        sa.select(models.inventory_items.c.id, models.inventory_items.c.sku, models.inventory_items.c.base_unit_id).where(
            models.inventory_items.c.organization_id == organization_id
        )
    ).fetchall()
    item_map: dict[str, tuple[str, str]] = {}
    for it in all_items:
        clean_sku = it.sku.strip()
        item_map[clean_sku] = (it.id, it.base_unit_id)
        if clean_sku.isdigit():
            item_map[str(int(clean_sku))] = (it.id, it.base_unit_id)
            item_map[clean_sku.zfill(6)] = (it.id, it.base_unit_id)

    pza_unit = session.execute(
        sa.select(models.inventory_units.c.id).where(
            models.inventory_units.c.organization_id == organization_id,
            models.inventory_units.c.code.in_(["PZ", "PZA", "PIEZA"]),
        )
    ).scalars().first()
    if not pza_unit:
        pza_unit = session.execute(
            sa.select(models.inventory_units.c.id).where(
                models.inventory_units.c.organization_id == organization_id
            )
        ).scalars().first()

    # Step 4: Seed recipes
    configured_recipes: list[dict[str, Any]] = []
    skipped_recipes: list[dict[str, Any]] = []

    for r in parsed_recipes:
        r_sku = r["sku"].strip()
        r_sku_clean = r_sku.lstrip("'")

        # Skip modifier group 25xxx (milk/sweetener sub-recipes)
        if r_sku_clean.startswith("25") and len(r_sku_clean) == 5:
            skipped_recipes.append({
                "sku": r_sku,
                "name": r["name"],
                "group": r.get("group", "GENERAL"),
                "reason": "Modificador de leche/endulzante gestionado vía modifier_options, no como producto directo de venta.",
            })
            continue

        product_id = prod_map.get(r_sku_clean) or prod_map.get(r_sku_clean.zfill(5)) or (prod_map.get(str(int(r_sku_clean))) if r_sku_clean.isdigit() else None)
        if not product_id:
            skipped_recipes.append({
                "sku": r_sku,
                "name": r["name"],
                "group": r.get("group", "GENERAL"),
                "reason": f"Producto con SKU '{r_sku}' no encontrado en el catálogo de productos.",
            })
            continue

        # Extract and validate components
        valid_components: list[dict[str, Any]] = []
        has_invalid_insumo = False
        calculated_cost = Decimal("0")

        for c in r["components"]:
            try:
                qty = Decimal(str(c["quantity"] or "0"))
            except Exception:
                qty = Decimal("0")
            if qty <= 0:
                continue

            c_sku = c["insumo_sku"].strip()
            c_sku_clean = str(int(c_sku)) if c_sku.isdigit() else c_sku
            item_info = item_map.get(c_sku_clean) or item_map.get(c_sku)
            if not item_info:
                skipped_recipes.append({
                    "sku": r_sku,
                    "name": r["name"],
                    "group": r.get("group", "GENERAL"),
                    "reason": f"Insumo con clave '{c_sku}' ({c['insumo_name']}) no existe en el catálogo de insumos.",
                })
                has_invalid_insumo = True
                break

            try:
                u_cost = Decimal(str(c.get("unit_cost") or "0"))
                calculated_cost += qty * u_cost
            except Exception:
                pass

            valid_components.append({
                "item_id": item_info[0],
                "unit_id": item_info[1],
                "net_quantity": qty,
                "gross_quantity": qty,
                "waste_rate": Decimal("0"),
                "notes": None,
            })

        if has_invalid_insumo:
            continue

        if not valid_components:
            skipped_recipes.append({
                "sku": r_sku,
                "name": r["name"],
                "group": r.get("group", "GENERAL"),
                "reason": "La receta no contiene ingredientes activos (cantidades en 0).",
            })
            continue

        # Upsert recipe
        recipe_id = f"rec-prod-{r_sku_clean.lower()}"
        existing_rec = session.execute(
            sa.select(models.recipes.c.id).where(
                models.recipes.c.organization_id == organization_id,
                models.recipes.c.product_id == product_id,
            )
        ).scalars().first()

        if existing_rec:
            session.execute(
                models.recipe_components.delete().where(
                    models.recipe_components.c.recipe_id == existing_rec
                )
            )
            session.execute(
                models.recipes.delete().where(
                    models.recipes.c.id == existing_rec
                )
            )

        session.execute(
            models.recipes.insert().values(
                id=recipe_id,
                organization_id=organization_id,
                product_id=product_id,
                output_item_id=None,
                branch_id=None,
                recipe_type="sale",
                version=1,
                status="active",
                yield_quantity=Decimal("1"),
                yield_unit_id=pza_unit,
                valid_from=now,
                valid_to=None,
                created_at=now,
                updated_at=now,
            )
        )

        for idx, comp in enumerate(valid_components):
            session.execute(
                models.recipe_components.insert().values(
                    recipe_id=recipe_id,
                    item_id=comp["item_id"],
                    quantity_base_units=comp["gross_quantity"],
                    unit_id=comp["unit_id"],
                    net_quantity=comp["net_quantity"],
                    waste_rate=comp["waste_rate"],
                    gross_quantity=comp["gross_quantity"],
                    sort_order=idx + 1,
                    notes=None,
                )
            )

        configured_recipes.append({
            "sku": r_sku,
            "name": r["name"],
            "group": r.get("group", "GENERAL"),
            "components_count": len(valid_components),
            "theoretical_cost": float(calculated_cost.quantize(Decimal("0.0001"))),
            "price": float(r["price"]) if r.get("price") else None,
        })

    session.commit()

    return {
        "total_pdf_recipes": len(parsed_recipes),
        "configured_count": len(configured_recipes),
        "skipped_count": len(skipped_recipes),
        "configured_recipes": configured_recipes,
        "skipped_recipes": skipped_recipes,
    }


if __name__ == "__main__":
    from pathlib import Path
    from restaurant_os.database import get_engine

    print("Iniciando carga de recetas a la base de datos configurada...")
    
    # Locate project root
    root_candidates = [
        Path(__file__).resolve().parents[3],
        Path.cwd(),
        Path(__file__).resolve().parents[2],
        Path("/app"),
        Path("/app/apps/api"),
    ]
    pdf_path = "productosestructura.frx.pdf"
    excel_dir = "."
    for r in root_candidates:
        cand_pdf = r / "productosestructura.frx.pdf"
        if cand_pdf.exists():
            pdf_path = str(cand_pdf)
            excel_dir = str(r)
            break
        cand_xls = r / "PRODUCTOS.XLS"
        if cand_xls.exists():
            excel_dir = str(r)

    try:
        engine = get_engine()
        with Session(engine) as session:
            result = load_recipes_from_pdf(session, pdf_path=pdf_path, excel_dir=excel_dir)
            print("=" * 60)
            print(f"ÉXITO: Se configuraron {result['configured_count']} recetas en la base de datos.")
            print(f"Omitidas (modificadores): {result['skipped_count']} recetas.")
            print("=" * 60)
    except Exception as e:
        print(f"Error al conectar o cargar en la base de datos: {e}")
        print("Verifica que tu variable DATABASE_URL o RESTAURANTOS_DATABASE_URL esté configurada correctamente.")

