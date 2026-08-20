from __future__ import annotations

import io
from datetime import datetime, timezone

import openpyxl
from test_platform_api import (
    BRANCH_ID,
    _admin_headers,
    _branch_admin_fixture,
    _client_with_seeded_database,
    _login_headers,
    _open_shift,
)


def test_daily_reconciliation_calculation_and_balance():
    """Verify PRD-FR-221 / BDD-SC-343..347:
    Daily reconciliation consolidates initial cash, sales by payment method,
    direct purchase expenses, fixed expenses, cash withdrawals, and calculates
    theoretical expected cash and cash overage/shortage (sobrante/faltante).
    """
    client = _client_with_seeded_database()
    headers = _admin_headers()
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # 1. Open shift with $2,000 MXN (200,000 cents)
    shift_res = _open_shift(client, opening_cash_cents=200000, headers=headers)
    assert shift_res.status_code == 200

    # 2. Query daily reconciliation report for today
    res = client.get(
        f"/api/v1/reports/branch-reconciliation/daily?branch_id={BRANCH_ID}&date={today_str}",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    data = res.json()

    assert data["branch_id"] == BRANCH_ID
    assert data["date"] == today_str
    assert "balance" in data
    assert "initial_cash" in data["balance"]
    assert "total_sales_with_tax" in data["balance"]
    assert "card_payments" in data["balance"]
    assert "transfer_payments" in data["balance"]
    assert "credit_sales" in data["balance"]
    assert "supplier_expenses" in data["balance"]
    assert "fixed_expenses" in data["balance"]
    assert "cash_withdrawals" in data["balance"]
    assert "cash_deposits" in data["balance"]
    assert "expected_cash_in_register" in data["balance"]
    assert "physical_cash_count" in data["balance"]
    assert "difference" in data["balance"]

    # Tables breakdown
    assert "suppliers_breakdown" in data
    assert "fixed_expenses_breakdown" in data
    assert "transfers_breakdown" in data
    assert "credit_clients_breakdown" in data
    assert "withdrawals_breakdown" in data
    assert "audit" in data
    assert "reviewed" in data["audit"]


def test_multi_branch_consolidated_report():
    """Verify PRD-FR-222 / BDD-SC-348..352:
    Multi-branch consolidated report aggregates expenses by supplier,
    expenses by fixed type, and totals across branches for daily, weekly or monthly ranges.
    """
    client = _client_with_seeded_database()
    headers = _admin_headers()
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    res = client.get(
        f"/api/v1/reports/branch-reconciliation/consolidated?date_from={today_str}&date_to={today_str}",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    data = res.json()

    assert "date_from" in data
    assert "date_to" in data
    assert "branches" in data
    assert "supplier_totals" in data
    assert "fixed_expense_totals" in data
    assert "summary" in data


def test_reconciliation_audit_status_update():
    """Verify PRD-FR-222 / BDD-SC-348..352:
    Auditor / Manager can mark a branch daily reconciliation as reviewed
    with audit notes and persist in DB.
    """
    client = _client_with_seeded_database()
    headers = _admin_headers()
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    res = client.post(
        "/api/v1/reports/branch-reconciliation/audit",
        json={
            "branch_id": BRANCH_ID,
            "date": today_str,
            "reviewed": True,
            "notes": "Comprobantes físicos validados contra depósito bancario.",
        },
        headers=headers,
    )
    assert res.status_code == 200, res.text
    audit_data = res.json()
    assert audit_data["reviewed"] is True
    assert audit_data["notes"] == "Comprobantes físicos validados contra depósito bancario."
    assert audit_data["audited_by_user_id"] is not None

    # Verify that querying the report returns the persisted audit state
    rep = client.get(
        f"/api/v1/reports/branch-reconciliation/daily?branch_id={BRANCH_ID}&date={today_str}",
        headers=headers,
    )
    assert rep.status_code == 200
    assert rep.json()["audit"]["reviewed"] is True
    assert (
        rep.json()["audit"]["notes"]
        == "Comprobantes físicos validados contra depósito bancario."
    )


def test_reconciliation_excel_export():
    """Verify PRD-FR-222:
    Exports daily/monthly branch reconciliation workbook as standard .xlsx stream.
    """
    client = _client_with_seeded_database()
    headers = _admin_headers()
    today = datetime.now(timezone.utc)

    res = client.get(
        f"/api/v1/reports/branch-reconciliation/export?branch_id={BRANCH_ID}&month={today.month}&year={today.year}",
        headers=headers,
    )
    assert res.status_code == 200, res.text
    assert (
        res.headers.get("content-type")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook_bytes = io.BytesIO(res.content)
    wb = openpyxl.load_workbook(workbook_bytes)
    assert "Master" in wb.sheetnames or "Resumen" in wb.sheetnames or "1" in wb.sheetnames


def test_reconciliation_rbac_enforcement():
    """Verify PRD-NFR-006 / PRD-NFR-020:
    Unauthorized or out-of-scope users cannot read or export financial reports.
    """
    client = _client_with_seeded_database()
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # 1. Unauthenticated requests are rejected
    unauth = client.get(
        f"/api/v1/reports/branch-reconciliation/daily?branch_id={BRANCH_ID}&date={today_str}"
    )
    assert unauth.status_code == 401

    # 2. Supervisor of Branch B cannot access Branch A
    fixture = _branch_admin_fixture(client)
    supervisor_headers = _login_headers(
        client, "supervisor.norte@kiwi.local", "Temporal123+"
    )
    forbidden = client.get(
        f"/api/v1/reports/branch-reconciliation/daily?branch_id={BRANCH_ID}&date={today_str}",
        headers=supervisor_headers,
    )
    assert forbidden.status_code == 403

    # 3. Supervisor can access their assigned branch
    allowed = client.get(
        f"/api/v1/reports/branch-reconciliation/daily?branch_id={fixture['branch_id']}&date={today_str}",
        headers=supervisor_headers,
    )
    assert allowed.status_code == 200


def test_supervisor_step_up_authorization():
    """Verify PRD-NFR-019 / FIX-05:
    Validates supervisor authorization step-up endpoint with PIN or password.
    """
    client = _client_with_seeded_database()
    fixture = _branch_admin_fixture(client)

    # Assign employee code SUP777 to supervisor
    assigned = client.put(
        f"/api/v1/users/{fixture['supervisor_id']}",
        headers=_admin_headers(),
        json={"employee_code": "SUP777"},
    )
    assert assigned.status_code == 200

    # 1. Valid PIN in supervisor's branch succeeds
    auth_res = client.post(
        "/api/v1/auth/supervisor-authorize",
        json={
            "supervisor_pin": "SUP777",
            "branch_id": fixture["branch_id"],
            "permission_code": "orders.discount.authorize",
        },
        headers=_admin_headers(),
    )
    assert auth_res.status_code == 200
    assert auth_res.json()["authorized"] is True
    assert auth_res.json()["supervisor_user_id"] == fixture["supervisor_id"]

    # 2. Invalid PIN fails
    invalid_res = client.post(
        "/api/v1/auth/supervisor-authorize",
        json={
            "supervisor_pin": "WRONG0",
            "branch_id": fixture["branch_id"],
        },
        headers=_admin_headers(),
    )
    assert invalid_res.status_code == 403
